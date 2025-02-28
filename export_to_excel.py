import json
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ✅ Load configuration from config.json
with open("config.json", "r") as config_file:
    config = json.load(config_file)

database = config["database_file"]  # Corrected key name
output_excel = config["excel_file"]  # Matches the config.json structure

# ✅ Connect to SQLite database
conn = sqlite3.connect(database)

# ✅ Retrieve stock data
query = "SELECT * FROM stock_data ORDER BY Date"
df = pd.read_sql(query, conn)

# ✅ Close database connection
conn.close()

# ✅ Create an Excel writer
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    # ✅ Create an overview sheet
    overview_df = pd.DataFrame()
    stocks = df["Stock_Symbol"].unique()

    for stock in stocks:
        stock_df = df[df["Stock_Symbol"] == stock].drop(columns=["Stock_Symbol"]).reset_index(drop=True)
        stock_df.to_excel(writer, sheet_name=stock, index=False)  # Save individual stock sheets

        # ✅ Append stock data to overview
        if overview_df.empty:
            overview_df = stock_df
        else:
            overview_df = pd.concat([overview_df, pd.DataFrame(columns=[""]), stock_df], axis=1)

    # ✅ Save the overview sheet
    overview_df.to_excel(writer, sheet_name="DataOverview", index=False)

# ✅ Apply alternating row colors for better readability
wb = writer.book
for sheet in wb.sheetnames:
    ws = wb[sheet]
    fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")

    for row in range(2, ws.max_row + 1, 2):  # Apply fill to even rows
        for cell in ws[row]:
            cell.fill = fill

wb.save(output_excel)
print(f"✅ Stock data successfully exported to {output_excel}")
